import  json
from    openpyxl                import load_workbook
from    datetime                import datetime, timedelta, timezone
from    previsao_trens.models   import Trem, Restricao, TremVazio
import  pandas as pd
from    django.conf import settings
import  os

tipo_desc_linhas = {

            'TGG_RUMO_FARELO':  8,
            'TGG_MRS_FARELO':   9,
            'TGG_VLI_FARELO':   10,
            'TGG_RUMO_MILHO':   11,
            'TGG_MRS_MILHO':    12,
            'TGG_VLI_MILHO':    13,
            'TGG_RUMO_SOJA':    14,
            'TGG_MRS_SOJA':     15,
            'TGG_VLI_SOJA':     16,

            'TEG_RUMO_FARELO':18,
            'TEG_MRS_FARELO':19,
            'TEG_VLI_FARELO':20,
            'TEG_RUMO_MILHO':21,
            'TEG_MRS_MILHO':22,
            'TEG_VLI_MILHO':23,
            'TEG_RUMO_SOJA':24,
            'TEG_MRS_SOJA':25,
            'TEG_VLI_SOJA':26,

            'TEAG_RUMO_FARELO':28,
            'TEAG_MRS_FARELO':29,
            'TEAG_VLI_FARELO':30,
            'TEAG_RUMO_MILHO':31,
            'TEAG_MRS_MILHO':32,
            'TEAG_VLI_MILHO':33,
            'TEAG_RUMO_SOJA':34,
            'TEAG_MRS_SOJA':35,
            'TEAG_VLI_SOJA':36,
            'TEAG_RUMO_AÇÚCAR':37,
            'TEAG_MRS_AÇÚCAR':38,
            'TEAG_VLI_AÇÚCAR':39,

            'CUTRALE_RUMO_FARELO':41,
            'CUTRALE_MRS_FARELO':42,
            'CUTRALE_VLI_FARELO':43,
            'CUTRALE_RUMO_MILHO':44,
            'CUTRALE_MRS_MILHO':45,
            'CUTRALE_VLI_MILHO':46,
            'CUTRALE_RUMO_SOJA':47,
            'CUTRALE_MRS_SOJA':48,
            'CUTRALE_VLI_SOJA':49,

            'TIPLAM_RUMO_FARELO':51,
            'TIPLAM_MRS_FARELO':52,
            'TIPLAM_VLI_FARELO':53,
            'TIPLAM_RUMO_MILHO':54,
            'TIPLAM_MRS_MILHO':55,
            'TIPLAM_VLI_MILHO':56,
            'TIPLAM_RUMO_SOJA':57,
            'TIPLAM_MRS_SOJA':58,
            'TIPLAM_VLI_SOJA':59,
            'TIPLAM_RUMO_AÇÚCAR':60,
            'TIPLAM_MRS_AÇÚCAR':61,
            'TIPLAM_VLI_AÇÚCAR':62,

            'T39_RUMO_FARELO':64,
            'T39_MRS_FARELO':65,
            'T39_VLI_FARELO':66,
            'T39_RUMO_MILHO':67,
            'T39_MRS_MILHO':68,
            'T39_VLI_MILHO':69,
            'T39_RUMO_SOJA':70,
            'T39_MRS_SOJA':71,
            'T39_VLI_SOJA':72,
            'T39_RUMO_AÇÚCAR':73,
            'T39_MRS_AÇÚCAR':74,
            'T39_VLI_AÇÚCAR':75,

            'TES_RUMO_FARELO':77,
            'TES_MRS_FARELO':78,
            'TES_VLI_FARELO':79,
            'TES_RUMO_MILHO':80,
            'TES_MRS_MILHO':81,
            'TES_VLI_MILHO':82,
            'TES_RUMO_SOJA':83,
            'TES_MRS_SOJA':84,
            'TES_VLI_SOJA':85,
            'TES_RUMO_AÇÚCAR':86,
            'TES_MRS_AÇÚCAR':87,
            'TES_VLI_AÇÚCAR':88,

            'MOEGA V_RUMO_FARELO':90,
            'MOEGA V_MRS_FARELO':91,
            'MOEGA V_VLI_FARELO':92,
            'MOEGA V_RUMO_MILHO':93,
            'MOEGA V_MRS_MILHO':94,
            'MOEGA V_VLI_MILHO':95,
            'MOEGA V_RUMO_SOJA':96,
            'MOEGA V_MRS_SOJA':97,
            'MOEGA V_VLI_SOJA':98,
            'MOEGA V_RUMO_AÇÚCAR':99,
            'MOEGA V_MRS_AÇÚCAR':100,
            'MOEGA V_VLI_AÇÚCAR':101,

            'MOEGA X_RUMO_FARELO':103,
            'MOEGA X_MRS_FARELO':104,
            'MOEGA X_VLI_FARELO':105,
            'MOEGA X_RUMO_MILHO':106,
            'MOEGA X_MRS_MILHO':107,
            'MOEGA X_VLI_MILHO':108,
            'MOEGA X_RUMO_SOJA':109,
            'MOEGA X_MRS_SOJA':110,
            'MOEGA X_VLI_SOJA':111,
            'MOEGA X_RUMO_AÇÚCAR':112,
            'MOEGA X_MRS_AÇÚCAR':113,
            'MOEGA X_VLI_AÇÚCAR':114,

            'TGRAO_RUMO_FARELO':116,
            'TGRAO_MRS_FARELO':117,
            'TGRAO_VLI_FARELO':118,
            'TGRAO_RUMO_MILHO':119,
            'TGRAO_MRS_MILHO':120,
            'TGRAO_VLI_MILHO':121,
            'TGRAO_RUMO_SOJA':122,
            'TGRAO_MRS_SOJA':123,
            'TGRAO_VLI_SOJA':124,
            'TGRAO_RUMO_AÇÚCAR':125,
            'TGRAO_MRS_AÇÚCAR':126,
            'TGRAO_VLI_AÇÚCAR':127,

            'CLI_RUMO_FARELO':129,
            'CLI_MRS_FARELO':130,
            'CLI_VLI_FARELO':131,
            'CLI_RUMO_MILHO':132,
            'CLI_MRS_MILHO':133,
            'CLI_VLI_MILHO':134,
            'CLI_RUMO_SOJA':135,
            'CLI_MRS_SOJA':136,
            'CLI_VLI_SOJA':137,
            'CLI_RUMO_AÇÚCAR':138,
            'CLI_MRS_AÇÚCAR':139,
            'CLI_VLI_AÇÚCAR':140,

            'TAC_RUMO_FARELO':142,
            'TAC_MRS_FARELO':143,
            'TAC_VLI_FARELO':144,
            'TAC_RUMO_MILHO':145,
            'TAC_MRS_MILHO':146,
            'TAC_VLI_MILHO':147,
            'TAC_RUMO_SOJA':148,
            'TAC_MRS_SOJA':149,
            'TAC_VLI_SOJA':150,
            'TAC_RUMO_AÇÚCAR':151,
            'TAC_MRS_AÇÚCAR':152,
            'TAC_VLI_AÇÚCAR':153,

            'T12A_RUMO_FARELO':155,
            'T12A_MRS_FARELO':156,
            'T12A_VLI_FARELO':157,
            'T12A_RUMO_MILHO':158,
            'T12A_MRS_MILHO':159,
            'T12A_VLI_MILHO':160,
            'T12A_RUMO_SOJA':161,
            'T12A_MRS_SOJA':162,
            'T12A_VLI_SOJA':163,
            'T12A_RUMO_AÇÚCAR':164,
            'T12A_MRS_AÇÚCAR':165,
            'T12A_VLI_AÇÚCAR':166,

            'BRACELL_RUMO_CELULOSE':168,
            'BRACELL_MRS_CELULOSE':169,
            'BRACELL_VLI_CELULOSE':170,

            'ELDORADO_RUMO_CELULOSE':172,
            'ELDORADO_MRS_CELULOSE':173,
            'ELDORADO_VLI_CELULOSE':174,

            'DPW_RUMO_CELULOSE':176,
            'DPW_MRS_CELULOSE':177,
            'DPW_VLI_CELULOSE':178,

            'TERLOC_RUMO_BRADO':180,
            'TERLOC_MRS_BRADO':181,
            'TERLOC_VLI_BRADO':182,


        }

class EXPORTAR_PLANILHA():

    def __init__(self):

        self.PLANILHA           = load_workbook("previsao_trens/src/DICIONARIOS/planilha_planner.xlsx", keep_vba=False) 
        self.PERIODO_VIGENTE    = pd.read_csv(f"previsao_trens/src/PARAMETROS/PERIODO_VIGENTE.csv", sep=";", index_col=0)
        
    def atualizar_a_data(self):

        LINHA       = self.PERIODO_VIGENTE[self.PERIODO_VIGENTE['NM_DIA'] == "D"]
        DATA_ARQ    = LINHA['DATA_ARQ'].values[0]
        DATETIME_D  = datetime.strptime(DATA_ARQ, "%Y-%m-%d")
        
        PREVISAO = self.PLANILHA['PREVISAO']    
        PREVISAO['X5'].value = DATETIME_D

    def inserir_previsao(self):

        CRIAR_TREM  = self.PLANILHA['PREVISAO']
        QUERYSET    = Trem.objects.all().order_by('previsao')
        
        if QUERYSET.exists():
                    
            for j, TREM in enumerate(QUERYSET):

                j = j + 5 # É PQ COMEÇA NA LINHA 5

                CRIAR_TREM.cell(row = j, column=2 ,   value=j - 5)
                CRIAR_TREM.cell(row = j, column=3 ,   value=TREM.prefixo)
                CRIAR_TREM.cell(row = j, column=4 ,   value=TREM.os)
                CRIAR_TREM.cell(row = j, column=5 ,   value=TREM.origem)
                CRIAR_TREM.cell(row = j, column=6 ,   value=TREM.destino)              
                CRIAR_TREM.cell(row = j, column=7 ,   value=TREM.terminal.nome)
                CRIAR_TREM.cell(row = j, column=8 ,   value=TREM.vagoes)
                CRIAR_TREM.cell(row = j, column=9 ,   value=TREM.mercadoria.nome)
                CRIAR_TREM.cell(row = j, column=10,   value=TREM.previsao.replace(tzinfo=None))
                CRIAR_TREM.cell(row = j, column=11,   value=TREM.ferrovia)

    def inserir_navegacao(self):

        NAVEGACAO = self.PLANILHA['NAVEGACAO']  
        COLUNAS   = { "D": 5, "D+1": 35, "D+2": 65, "D+3": 95, "D+4": 125 }

        with open(f"previsao_trens/src/DICIONARIOS/MAPA_TERMINAIS_EXCEL.json") as ARQUIVO:
            map_TERMINAIS = json.load(ARQUIVO)

        PERIODO_VIGENTE    = pd.read_csv("previsao_trens/src/PARAMETROS/PERIODO_VIGENTE.csv",   encoding='utf-8-sig', sep=';', index_col=0)
        TERMINAIS_ATIVOS   = pd.read_csv("previsao_trens/src/PARAMETROS/DESCARGAS_ATIVAS.csv",  encoding='utf-8-sig', sep=';', index_col=0)
        RESTRICOES_ATIVAS  = pd.read_csv("previsao_trens/src/PARAMETROS/RESTRICOES_ATIVAS.csv", encoding='utf-8-sig', sep=';', index_col=0)

        lsTERMINAIS_ATIVOS      = TERMINAIS_ATIVOS[TERMINAIS_ATIVOS['TERMINAL'] > 0].index.tolist()
        lsTERMINAIS_DESATIVADOS = TERMINAIS_ATIVOS[TERMINAIS_ATIVOS['TERMINAL'] == 0].index.tolist() 
        TERMINAIS_ATIVOS.drop('TERMINAL', axis=1, inplace=True)

        TERMINAIS_COM_RESTRICAO = RESTRICOES_ATIVAS[RESTRICOES_ATIVAS['RESTRICAO'] > 0].index.tolist()
        TERMINAIS_SEM_RESTRICAO = RESTRICOES_ATIVAS[RESTRICOES_ATIVAS['RESTRICAO'] == 0].index.tolist()

        for TERMINAL in lsTERMINAIS_ATIVOS:
            
            if TERMINAL in map_TERMINAIS:

                #region OCULTANDO RESTRICAO SE NÃO HOUVER 
                
                if TERMINAL in TERMINAIS_SEM_RESTRICAO:
                    
                    LINHA_PCT    = map_TERMINAIS[TERMINAL]["LIMITES"]["FIM"]
                    LINHA_MOTIVO = map_TERMINAIS[TERMINAL]["LIMITES"]["FIM"] - 1
                    
                    NAVEGACAO.row_dimensions[LINHA_PCT].hidden = True 
                    NAVEGACAO.row_dimensions[LINHA_MOTIVO].hidden = True 


               
                #endregion

                DESCARGAS_ATIVAS = TERMINAIS_ATIVOS.loc[TERMINAL][TERMINAIS_ATIVOS.loc[TERMINAL] > 0].index.tolist()
                DESCARGAS_ATIVAS = [item.split('_') for item in DESCARGAS_ATIVAS] #  <-- [['RUMO', 'FARELO'], ['RUMO', 'SOJA'], ['MRS', 'SOJA'], ['RUMO', 'MILHO']]  

                LINHA_PREFIXO = map_TERMINAIS[TERMINAL]["LIMITES"]["INICIO"] + 2
                LINHA_VAGOES  = map_TERMINAIS[TERMINAL]["LIMITES"]["INICIO"] + 3
                LINHA_PEDRA   = map_TERMINAIS[TERMINAL]["LIMITES"]["INICIO"] + 4
                LINHA_FINAL   = map_TERMINAIS[TERMINAL]["LIMITES"]["FIM"] - 2
                
                #region INSERINDO VALORES
                for DIA in COLUNAS.keys():

                    DATA_ARQ = PERIODO_VIGENTE[PERIODO_VIGENTE['NM_DIA'] == DIA].iloc[0]['DATA_ARQ']

                    with open(f"previsao_trens/src/DESCARGAS/{ TERMINAL }/descarga_{ DATA_ARQ }.json") as ARQUIVO_DESCARGA:
                        DESCARGA = json.load(ARQUIVO_DESCARGA)

                    #region INSERINDO PREFIXO E CHEGADA
                    for i in range(24):
                        
                        NAVEGACAO.cell(row = LINHA_PREFIXO, column=(i + COLUNAS[DIA]),   value=DESCARGA["PREFIXO"][i][0])
                        NAVEGACAO.cell(row = LINHA_VAGOES,  column=(i + COLUNAS[DIA]),   value=DESCARGA["CHEGADA"][i][0])
                        NAVEGACAO.cell(row = LINHA_PEDRA,   column=(i + COLUNAS[DIA]),   value=DESCARGA["PEDRA"][i])
                    
                    #endregion

                    #region INSERINDO DESCARGAS
                    for DESCARGA_ATIVA in DESCARGAS_ATIVAS:
                        
                        FERROVIA = DESCARGA_ATIVA[0]
                        PRODUTO  = DESCARGA_ATIVA[1] 

                        LINHA_ENCOSTE       = map_TERMINAIS[TERMINAL][FERROVIA][PRODUTO]
                        LINHA_PRODUTIVIDADE = LINHA_ENCOSTE + 2
                        LINHA_SALDO         = LINHA_ENCOSTE + 1

                        if DIA == "D":
                            SALDO_VIRADA = DESCARGA["DESCARGAS"][FERROVIA][PRODUTO]["INDICADORES"]["SALDO_DE_VIRADA"]
                            NAVEGACAO.cell(row =LINHA_ENCOSTE + 1, column=4,   value=SALDO_VIRADA)

                        for i in range(24):
                        
                            NAVEGACAO.cell(row = LINHA_ENCOSTE,          column=(i + COLUNAS[DIA]),   value=DESCARGA["DESCARGAS"][FERROVIA][PRODUTO]["ENCOSTE"][i][0])
                            NAVEGACAO.cell(row = LINHA_SALDO,            column=(i + COLUNAS[DIA]),   value=DESCARGA["DESCARGAS"][FERROVIA][PRODUTO]["SALDO"][i])
                            NAVEGACAO.cell(row = LINHA_PRODUTIVIDADE,    column=(i + COLUNAS[DIA]),   value=DESCARGA["DESCARGAS"][FERROVIA][PRODUTO]["PRODUTIVIDADE"][i])

                        #region INSERINDO TOTAIS PEDRA CHEGADA

                        NAVEGACAO.cell(row = LINHA_SALDO,            column=(24 + COLUNAS[DIA]),   value=DESCARGA["DESCARGAS"][FERROVIA][PRODUTO]["INDICADORES"]["TOTAL_CHEGADA"])
                        NAVEGACAO.cell(row = LINHA_PRODUTIVIDADE,    column=(25 + COLUNAS[DIA]),   value=DESCARGA["DESCARGAS"][FERROVIA][PRODUTO]["INDICADORES"]["TOTAL_PRODUTIVIDADE"])    

                        #endregion            

                    #endregion
                    
                    #region inserindo totais

                    #INSERINDO PEDRA


                    NAVEGACAO.cell(row = LINHA_PEDRA,   column=(24 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["TOTAL_SALDO"])
                    NAVEGACAO.cell(row = LINHA_PEDRA,   column=(25 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["TOTAL_PEDRA"])


                    #TOTAIS POR PERIODO DAS FERROVIAS
                    NAVEGACAO.cell(row = LINHA_FINAL,   column=( 2 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["PEDRAS"]["RUMO"]["P1"])
                    NAVEGACAO.cell(row = LINHA_FINAL,   column=( 3 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["PEDRAS"]["MRS"]["P1"])
                    NAVEGACAO.cell(row = LINHA_FINAL,   column=( 4 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["PEDRAS"]["VLI"]["P1"])
                    NAVEGACAO.cell(row = LINHA_FINAL,   column=( 5 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["PEDRAS"]["TOTAL"]["P1"])

                    NAVEGACAO.cell(row = LINHA_FINAL,   column=( 8 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["PEDRAS"]["RUMO"]["P2"])
                    NAVEGACAO.cell(row = LINHA_FINAL,   column=( 9 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["PEDRAS"]["MRS"]["P2"])
                    NAVEGACAO.cell(row = LINHA_FINAL,   column=(10 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["PEDRAS"]["VLI"]["P2"])
                    NAVEGACAO.cell(row = LINHA_FINAL,   column=(11 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["PEDRAS"]["TOTAL"]["P2"])

                    NAVEGACAO.cell(row = LINHA_FINAL,   column=(14 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["PEDRAS"]["RUMO"]["P3"])
                    NAVEGACAO.cell(row = LINHA_FINAL,   column=(15 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["PEDRAS"]["MRS"]["P3"])
                    NAVEGACAO.cell(row = LINHA_FINAL,   column=(16 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["PEDRAS"]["VLI"]["P3"])
                    NAVEGACAO.cell(row = LINHA_FINAL,   column=(17 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["PEDRAS"]["TOTAL"]["P3"])

                    NAVEGACAO.cell(row = LINHA_FINAL,   column=(20 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["PEDRAS"]["RUMO"]["P4"])
                    NAVEGACAO.cell(row = LINHA_FINAL,   column=(21 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["PEDRAS"]["MRS"]["P4"])
                    NAVEGACAO.cell(row = LINHA_FINAL,   column=(22 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["PEDRAS"]["VLI"]["P4"])
                    NAVEGACAO.cell(row = LINHA_FINAL,   column=(23 + COLUNAS[DIA]),   value=DESCARGA["INDICADORES"]["PEDRAS"]["TOTAL"]["P4"])


                    #endregion



                    if TERMINAL in TERMINAIS_COM_RESTRICAO:

                        DESCARGAS_RESTRICOES    = RESTRICOES_ATIVAS.drop('RESTRICAO', axis=1)
                        PRODUTO_RESTRICAO       = DESCARGAS_RESTRICOES.loc[TERMINAL][DESCARGAS_RESTRICOES.loc[TERMINAL] > 0].index.tolist()

                        #É PARA PEGAR UMA RESTRICAO
                        FERROVIA_ALEATORA_DO_TERMINAL = next(iter(DESCARGA["DESCARGAS"]))

                        DESCARGA["RESTRICAO_PCT"] = DESCARGA["DESCARGAS"][FERROVIA_ALEATORA_DO_TERMINAL][PRODUTO_RESTRICAO[0]]["RESTRICAO"]

                        LINHA_PCT    = map_TERMINAIS[TERMINAL]["LIMITES"]["FIM"]
                        LINHA_MOTIVO = map_TERMINAIS[TERMINAL]["LIMITES"]["FIM"] - 1

                        for i in range(24):
                            if not  DESCARGA["RESTRICAO_PCT"][i] == 0:
                                NAVEGACAO.cell(row =LINHA_PCT,     column=(i + COLUNAS[DIA]),  value=DESCARGA["RESTRICAO_PCT"][i])
                                NAVEGACAO.cell(row =LINHA_MOTIVO,  column=(i + COLUNAS[DIA]),  value=DESCARGA["RESTRICAO_MOTIVO"][i])

                #endregion   

                #region OCULTANDO DESCARDAS DESATIVADAS

                DESCARGAS_DESATIVADAS = TERMINAIS_ATIVOS.loc[TERMINAL][TERMINAIS_ATIVOS.loc[TERMINAL] == 0].index.tolist()
                DESCARGAS_DESATIVADAS = [item.split('_') for item in DESCARGAS_DESATIVADAS] #  <-- [['RUMO', 'FARELO'], ['RUMO', 'SOJA'], ['MRS', 'SOJA'], ['RUMO', 'MILHO']]  
    
                for DESCARGA_DESATIVA in DESCARGAS_DESATIVADAS:

                    FERROVIA = DESCARGA_DESATIVA[0]
                    PRODUTO  = DESCARGA_DESATIVA[1] 

                    if FERROVIA in map_TERMINAIS[TERMINAL] and PRODUTO in map_TERMINAIS[TERMINAL][FERROVIA]:

                        LINHA_ENCOSTE       = map_TERMINAIS[TERMINAL][FERROVIA][PRODUTO]
                        LINHA_SALDO         = map_TERMINAIS[TERMINAL][FERROVIA][PRODUTO]  + 1
                        LINHA_PRODUTIVIDADE = map_TERMINAIS[TERMINAL][FERROVIA][PRODUTO]  + 2

                        NAVEGACAO.row_dimensions[LINHA_ENCOSTE].hidden       = True
                        NAVEGACAO.row_dimensions[LINHA_SALDO].hidden         = True
                        NAVEGACAO.row_dimensions[LINHA_PRODUTIVIDADE].hidden = True
                #endregion

        #region REMOVENDO TERMINAIS DESATIVOS
        for TERMINAL in lsTERMINAIS_DESATIVADOS:
            if TERMINAL in map_TERMINAIS:
                TAMANHO_TERMINAL = map_TERMINAIS[TERMINAL]["LIMITES"]
                for i in range(TAMANHO_TERMINAL["INICIO"], TAMANHO_TERMINAL["FIM"] +2):
                   
                    NAVEGACAO.row_dimensions[i].hidden = True            
        #endregion

    def inserir_restricao(self):


        RESTRICOES = Restricao.objects.all()
        wsRESTRICAO = self.PLANILHA['RESTRICAO']   

        for i, RESTRICAO in enumerate(RESTRICOES):

            i = i + 5 # É PQ COMEÇA NA LINHA 5

            wsRESTRICAO.cell(row = i, column=2,   value=i-5)
            wsRESTRICAO.cell(row = i, column=3,   value=RESTRICAO.terminal)
            wsRESTRICAO.cell(row = i, column=4,   value=RESTRICAO.mercadoria)

            wsRESTRICAO.cell(row = i, column=5,   value=RESTRICAO.comeca_em.replace(tzinfo=None))
            wsRESTRICAO.cell(row = i, column=6,   value=RESTRICAO.termina_em.replace(tzinfo=None))

            wsRESTRICAO.cell(row = i, column=7,   value=RESTRICAO.porcentagem)
            wsRESTRICAO.cell(row = i, column=8,   value=RESTRICAO.motivo)
            wsRESTRICAO.cell(row = i, column=9,   value=RESTRICAO.comentario)

    def inserir_folha_capa(self, USUARIO_LOGADO):
        
        HOME          = self.PLANILHA['HOME']
        
        HOME.cell(row = 6, column=5,  value=f"{USUARIO_LOGADO.first_name} {USUARIO_LOGADO.last_name}" )
        HOME.cell(row = 7, column=5,  value=USUARIO_LOGADO.email )
        HOME.cell(row = 8, column=5,  value=datetime.today() - timedelta(hours=3))

    def inserir_previsao_subida(self):

        SUBIDA = self.PLANILHA['SUBIDA']

        QUERYSET = TremVazio.objects.all().order_by('previsao')

        for j, TREM in enumerate(QUERYSET):

            j = j + 4 # É PQ COMEÇA NA LINHA 5

            SUBIDA.cell(row = j, column=2,   value=j - 4)
            SUBIDA.cell(row = j, column=3,   value=TREM.prefixo)
            SUBIDA.cell(row = j, column=4,   value=TREM.margem)
            SUBIDA.cell(row = j, column=5,   value=TREM.ferrovia)
            SUBIDA.cell(row = j, column=6,   value=TREM.previsao.replace(tzinfo=None))
            SUBIDA.cell(row = j, column=7,   value=TREM.eot)

            SUBIDA.cell(row = j, column=8,    value=TREM.qt_graos)
            SUBIDA.cell(row = j, column=9,    value=TREM.qt_ferti)
            SUBIDA.cell(row = j, column=10,    value=TREM.qt_celul)
            SUBIDA.cell(row = j, column=11,   value=TREM.qt_acuca)
            SUBIDA.cell(row = j, column=12,   value=TREM.qt_contei)

            SUBIDA.cell(row = j, column=13,   value=TREM.loco_1)
            SUBIDA.cell(row = j, column=14,   value=TREM.loco_2)
            SUBIDA.cell(row = j, column=15,   value=TREM.loco_3)
            SUBIDA.cell(row = j, column=16,   value=TREM.loco_4)
            SUBIDA.cell(row = j, column=17,   value=TREM.loco_5)


    
    def salvar(self):
        
        caminho_static  = os.path.join(settings.STATIC_URL, 'downloads')
        caminho_arquivo = os.path.join(caminho_static, "PLANILHA_DESCARGA.xlsm")
        self.PLANILHA.save(caminho_arquivo)

def gerar_planilha(USUARIO_LOGADO):
    
    PLANILHA = EXPORTAR_PLANILHA()
    PLANILHA.inserir_previsao()
    PLANILHA.inserir_navegacao()
    PLANILHA.inserir_restricao()
    PLANILHA.inserir_previsao_subida()
    PLANILHA.inserir_folha_capa(USUARIO_LOGADO)
    
    return PLANILHA.PLANILHA
    

#############################################################################################################

class PlanilhaAntiga_old:

    def __init__(self):

        self.planilha        = load_workbook("previsao_trens/src/DICIONARIOS/planilha_antiga.xlsm", keep_vba=True)
        self.periodo_vigente = pd.read_csv(f"previsao_trens/src/PARAMETROS/PERIODO_VIGENTE.csv",    sep=";", index_col=0)
        
        self.periodo_vigente = self.periodo_vigente.drop(self.periodo_vigente.index[0]) 

    def atualizar_data(self):

        linha       = self.periodo_vigente[self.periodo_vigente["NM_DIA"] == "D"]
        data_arq    = linha["DATA_ARQ"].values[0]
        datetime_D  = datetime.strptime(data_arq, "%Y-%m-%d")
  
        ws_previsao = self.planilha["Previsão"]
        ws_previsao["X5"].value = datetime_D.replace(tzinfo=None)

    def inserir_previsao(self):

        produtos_substituir  = { "SOJA": "Soja", "ACUCAR": "Açúcar", "MILHO": "Milho", "FARELO": "Farelo", "CELULOSE": "Celulose" }
        terminais_substituir = { "CLI": "RUMO", "CLI ACUCAR": "RUMO", "COPERSUCAR": "TAC", "TAC ACUCAR": "TAC", "MOEGA X": "Moega X", "MOEGA V": "Moega V", "T39": "Moega IV", "TGRAO": "Tgrão", "TEAG ACUCAR": "TEAG" }
        segmento             = { "SOJA": "Grão", "FARELO": "Grão", "MILHO": "Grão", "ACUCAR": "Açúcar", "CELULOSE": "Industrial", "CONTEINER": "Industrial", "KCL": "Fertilizante", "UREIA": "Fertilizante", "OUTROS": "Fertilizante"}
        linhas               = { "D": 15, "D+1": 54, "D+2": 100, "D+3": 173, "D+4": 250 }
        

        ws_criarTrem = self.planilha["Criar_Trem"]

        for _, linha in self.periodo_vigente.iterrows():

            data = datetime.strptime(linha['DATA_ARQ'], '%Y-%m-%d')
            dia_logistico = linha['NM_DIA']

            queryset = Trem.objects.filter( previsao__year=data.year, previsao__month=data.month, previsao__day=data.day ).order_by('posicao_previsao')

            if queryset.exists():

                i = 0
                linha = linhas[dia_logistico]
                
                for trem in queryset:
                    
                    terminal = trem.terminal
                    if terminal in terminais_substituir: terminal = terminais_substituir[terminal]
                    
                    mercadoria = trem.mercadoria
                    if mercadoria in produtos_substituir: mercadoria = produtos_substituir[mercadoria]

                    ws_criarTrem.cell(row=linha + i, column=5,  value=trem.ferrovia)
                    ws_criarTrem.cell(row=linha + i, column=6,  value=trem.os)
                    ws_criarTrem.cell(row=linha + i, column=7,  value=trem.prefixo)
                    ws_criarTrem.cell(row=linha + i, column=8,  value=trem.origem)
                    ws_criarTrem.cell(row=linha + i, column=9,  value=trem.destino)
                    ws_criarTrem.cell(row=linha + i, column=10, value=trem.previsao.replace(tzinfo=None))

                    ws_criarTrem.cell(row=linha + i, column=13, value=terminal)     
                    ws_criarTrem.cell(row=linha + i, column=14, value=segmento[trem.mercadoria])    
                    ws_criarTrem.cell(row=linha + i, column=15, value=mercadoria)   
                    ws_criarTrem.cell(row=linha + i, column=16, value=trem.vagoes)
  
                    i = i + 1

    def inserir_produtividade(self):

        ws_navegacao = self.planilha['Navegação']
        colunas      = { "D": 51, "D+1": 86, "D+2": 121, "D+3": 156, "D+4": 191}
        linhas_pedra = { "TGG":  14, "TEG":  52, "TEAG": 90, "TEAG ACUCAR": 128, "CUTRALE": 148, "T39": 245, "TES": 283, "MOEGA V": 231,"MOEGA X": 350,"TGRAO":  397,"CLI": 345,"CLI ACUCAR":473,"TAC": 493, "TAC ACUCAR": 531,"T12A":551,"SUZANO": 590,"BRACELL": 607}

        df_descargas_ativas  = pd.read_csv("previsao_trens/src/PARAMETROS/DESCARGAS_ATIVAS.csv",  encoding='utf-8-sig', sep=';', index_col=0)
        ls_terminais_ativos  = df_descargas_ativas.index[df_descargas_ativas['TERMINAL'] != 0].tolist()

        for terminal in ls_terminais_ativos:

            if terminal in linhas_pedra:

                descargas_terminal = df_descargas_ativas.loc[terminal]
                descargas_ativas_terminal = [indice for indice, valor in descargas_terminal.items() if valor > 0]
                descargas_ativas_terminal.remove("TERMINAL")

                for tipo_desc in descargas_ativas_terminal:

                    linha_planinha = tipo_desc_linhas[f'{terminal}_{tipo_desc}']

                    ws_navegacao.row_dimensions[linhas_pedra[terminal]].hidden = False

                    ws_navegacao.row_dimensions[linha_planinha - 2].hidden = False
                    ws_navegacao.row_dimensions[linha_planinha - 1].hidden = False
                    ws_navegacao.row_dimensions[linha_planinha].hidden     = False

                    ferrovia = tipo_desc.split("_")[0]
                    produto  = tipo_desc.split("_")[1]

                    for _, linha in self.periodo_vigente.iterrows():

                        data_arq        = linha['DATA_ARQ']
                        dia_logistico   = linha['NM_DIA']
                        
                        
                        with open(f"previsao_trens/src/DESCARGAS/{ terminal }/descarga_{ data_arq }.json") as json_descarga:
                            descarga = json.load(json_descarga)

                        if dia_logistico == "D":

                            saldo_virada = descarga["DESCARGAS"][ferrovia][produto]["INDICADORES"]["SALDO_DE_VIRADA"]
                            ws_navegacao.cell(row=(linha_planinha - 1), column=15, value=saldo_virada)
                        
                        linha_produtividade = descarga["DESCARGAS"][ferrovia][produto]["PRODUTIVIDADE"]

                        for i, produtividade in enumerate(linha_produtividade):

                            ws_navegacao.cell(row=linha_planinha, column=colunas[dia_logistico] + i, value=produtividade)       

    def inserir_restricao(self):

        ws_restricao = self.planilha['Restrição']      

        primeira_linha = 6

        restricoes = Restricao.objects.all()
   
        terminais_substituir = {'CLI': 'RUMO', 'COPERSUCAR': 'TAC', 'MOEGA X': 'Moega X', 'MOEGA V': 'Moega V','T39': 'Moega IV','TAC ACUCAR': 'TAC','CLI ACUCAR': 'RUMO','TEAG ACUCAR': 'TEAG',}
        segmentos_substituir = {'SOJA': 'Grão', 'FARELO': 'Grão', 'MILHO': 'Grão', 'ACUCAR': 'Açúcar','CELULOSE': 'Indústrial'}
        
        for i, restricao in enumerate(restricoes):       

            terminal = restricao.terminal
            if terminal in terminais_substituir: terminal = terminais_substituir[terminal]

            segmento = restricao.mercadoria
            if segmento in segmentos_substituir: segmento = segmentos_substituir[segmento]

            ws_restricao.cell(row=(primeira_linha + i), column=13, value=terminal)
            ws_restricao.cell(row=(primeira_linha + i), column=14, value=segmento)
            ws_restricao.cell(row=(primeira_linha + i), column=15, value=restricao.comeca_em.replace(tzinfo=None))
            ws_restricao.cell(row=(primeira_linha + i), column=16, value=restricao.termina_em.replace(tzinfo=None))
            ws_restricao.cell(row=(primeira_linha + i), column=17, value=restricao.motivo)       
            ws_restricao.cell(row=(primeira_linha + i), column=18, value=restricao.porcentagem)

def gerar_planilha_antiga_old():

    Planilha = PlanilhaAntiga()
    Planilha.atualizar_data()
    Planilha.inserir_previsao()
    Planilha.inserir_produtividade()
    Planilha.inserir_restricao()

    return Planilha.planilha

#############################################################################################################

from    previsao_trens.packages.DETELHE.CARREGAR_PAGINA import CARREGAR_RELATORIO_DETALHE

class PlanilhaAntiga:

    def __init__(self):

            self.planilha        = load_workbook("previsao_trens/src/DICIONARIOS/planilha_antiga_clean.xlsx")
            self.periodo_vigente = pd.read_csv(f"previsao_trens/src/PARAMETROS/PERIODO_VIGENTE.csv",    sep=";", index_col=0)
            
            self.periodo_vigente = self.periodo_vigente.drop(self.periodo_vigente.index[0]) 

    def inserir_previsao(self):

            produtos_substituir  = { "SOJA": "Soja", "ACUCAR": "Açúcar", "MILHO": "Milho", "FARELO": "Farelo", "CELULOSE": "Celulose" }
            terminais_substituir = { "CLI": "RUMO", "CLI ACUCAR": "RUMO", "COPERSUCAR": "TAC", "TAC ACUCAR": "TAC", "MOEGA X": "Moega X", "MOEGA V": "Moega V", "T39": "Moega IV", "TGRAO": "Tgrão", "TEAG ACUCAR": "TEAG" }
            segmento             = { "SOJA": "Grão", "FARELO": "Grão", "MILHO": "Grão", "ACUCAR": "Açúcar", "CELULOSE": "Industrial", "CONTEINER": "Industrial", "KCL": "Fertilizante", "UREIA": "Fertilizante", "OUTROS": "Fertilizante"}
            linhas               = { "D": 15, "D+1": 54, "D+2": 100, "D+3": 173, "D+4": 250 }
            
            ws_criarTrem = self.planilha["Criar_Trem"]

            for _, linha in self.periodo_vigente.iterrows():

                data = datetime.strptime(linha['DATA_ARQ'], '%Y-%m-%d')
                dia_logistico = linha['NM_DIA']

                queryset = Trem.objects.filter(
                
                            previsao__year=data.year,
                            previsao__month=data.month,
                            previsao__day=data.day

                        ).order_by('posicao_previsao')
                
                if queryset.exists():

                    i = 0
                    linha = linhas[dia_logistico]
                    
                    for trem in queryset:
                        
                        terminal = trem.terminal.nome
                        if terminal in terminais_substituir: terminal = terminais_substituir[terminal]
                        
                        mercadoria = trem.mercadoria.nome
                        if mercadoria in produtos_substituir: mercadoria = produtos_substituir[mercadoria]

                        ws_criarTrem.cell(row=linha + i, column=5,  value=trem.ferrovia)
                        ws_criarTrem.cell(row=linha + i, column=6,  value=trem.os)
                        ws_criarTrem.cell(row=linha + i, column=7,  value=trem.prefixo)
                        ws_criarTrem.cell(row=linha + i, column=8,  value=trem.origem)
                        ws_criarTrem.cell(row=linha + i, column=9,  value=trem.destino)
                        ws_criarTrem.cell(row=linha + i, column=10, value=trem.previsao.replace(tzinfo=None))
                        ws_criarTrem.cell(row=linha + i, column=12, value=trem.encoste.replace(tzinfo=None))
                        

                        ws_criarTrem.cell(row=linha + i, column=13, value=terminal)     
                        ws_criarTrem.cell(row=linha + i, column=14, value=segmento[trem.mercadoria.nome])    
                        ws_criarTrem.cell(row=linha + i, column=15, value=mercadoria)   
                        ws_criarTrem.cell(row=linha + i, column=16, value=trem.vagoes)
    
                        i = i + 1

    def inserir_detalhe(self):

        ws_criarDetalhe = self.planilha["Detalhe"]

        COLUNAS = {
        0: {
            'SALDOS':       {'P1':  7, 'P2': 10, 'P3': 13, 'P4': 16},
            'RECEBIMENTOS': {'P1':  8, 'P2': 11, 'P3': 14, 'P4': 17}, 
            'PEDRA':        {'P1':  9, 'P2': 12, 'P3': 15, 'P4': 18}   
        },
        1: {
            'SALDOS':       {'P1': 21, 'P2': 24, 'P3': 27, 'P4': 30},
            'RECEBIMENTOS': {'P1': 22, 'P2': 25, 'P3': 28, 'P4': 31},
            'PEDRA':        {'P1': 23, 'P2': 26, 'P3': 29, 'P4': 32}
        },
        2: {
            'SALDOS':       {'P1': 35, 'P2': 38, 'P3': 41, 'P4': 44},
            'RECEBIMENTOS': {'P1': 36, 'P2': 39, 'P3': 42, 'P4': 45},
            'PEDRA':        {'P1': 37, 'P2': 40, 'P3': 43, 'P4': 46}
        },
        3: {
            'SALDOS':       {'P1': 49, 'P2': 52, 'P3': 55, 'P4': 58},
            'RECEBIMENTOS': {'P1': 50, 'P2': 53, 'P3': 56, 'P4': 59},
            'PEDRA':        {'P1': 51, 'P2': 54, 'P3': 57, 'P4': 60}
        },
        4: {
            'SALDOS':       {'P1': 63, 'P2': 66, 'P3': 69, 'P4': 72},
            'RECEBIMENTOS': {'P1': 64, 'P2': 67, 'P3': 70, 'P4': 73},
            'PEDRA':        {'P1': 65, 'P2': 68, 'P3': 71, 'P4': 74}
        },
    }

        with open(f"previsao_trens/src/DICIONARIOS/MAPA_DETALHE_PLANILHA_ANTIGA.json") as ARQUIVO:
            map_DETALHE = json.load(ARQUIVO)

        DADOS_RELATORIO = CARREGAR_RELATORIO_DETALHE()  # ['PRINCIPAL', 'RUMO', 'TOTAIS', 'VLI', 'MRS']
        DADOS_RELATORIO = DADOS_RELATORIO["PRINCIPAL"]  # ['TGG', 'TEG', 'CUTRALE', 'T39', 'TES', 'MOEGA X', 'TGRAO', 'CLI', 'TAC', 'T12A', 'BRACELL', 'SBR', 'ECOPORTO', 'HIDROVIAS', 'TEAG']

        for TERMINAL in DADOS_RELATORIO:
            if TERMINAL in map_DETALHE:    
                FERROVIAS = list(DADOS_RELATORIO[TERMINAL].keys())
                FERROVIAS.remove('MARGEM')
                FERROVIAS.remove('TOTAL') 

                for FERROVIA in FERROVIAS:
                    
                    PRODUTOS = list(DADOS_RELATORIO[TERMINAL][FERROVIA].keys()) 

                    for PRODUTO in PRODUTOS:
                        for DIA in [0, 1, 2, 3, 4]: 
                            for TITULO_COL in list(DADOS_RELATORIO[TERMINAL][FERROVIA][PRODUTO][DIA]):                       
                                if not TITULO_COL == "TT_OF" and  not TITULO_COL == "TT_PD":                             
                                    for PERIODO in DADOS_RELATORIO[TERMINAL][FERROVIA][PRODUTO][DIA][TITULO_COL]:

                                        LINHA   =   map_DETALHE[TERMINAL][FERROVIA][PRODUTO]
                                        COLUNA  =   COLUNAS[DIA][TITULO_COL][PERIODO] + 7
                                        VALOR   =   DADOS_RELATORIO[TERMINAL][FERROVIA][PRODUTO][DIA][TITULO_COL][PERIODO]

                                        ws_criarDetalhe.cell(row = LINHA, column= COLUNA,  value = VALOR )


def gerar_planilha_antiga():

    Planilha = PlanilhaAntiga()
    Planilha.inserir_previsao()
    Planilha.inserir_detalhe()

    return Planilha.planilha