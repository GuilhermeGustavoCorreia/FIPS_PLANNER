from    openpyxl                import load_workbook
from    previsao_trens.packages.DETELHE.CARREGAR_PAGINA import CARREGAR_RELATORIO_DETALHE
import  os
import  json
from    django.conf import settings


def BAIXAR_DETALHE():
    
    COLUNAS = {
        0: {
            'SALDOS':       {'P1':  7, 'P2': 10, 'P3': 13, 'P4': 16},
            'RECEBIMENTOS': {'P1':  8, 'P2': 11, 'P3': 14, 'P4': 17}, 
            'PEDRA':        {'P1':  9, 'P2': 12, 'P3': 15, 'P4': 18}   
        },
        1: {
            'SALDOS':       {'P1': 21, 'P2': 24, 'P3': 27, 'P4': 30},
            'RECEBIMENTOS': {'P1': 22, 'P2': 25, 'P3': 28, 'P4': 31},
            'PEDRA':        {'P1': 23, 'P2': 26, 'P3': 29, 'P4': 32}
        },
        2: {
            'SALDOS':       {'P1': 35, 'P2': 38, 'P3': 41, 'P4': 44},
            'RECEBIMENTOS': {'P1': 36, 'P2': 39, 'P3': 42, 'P4': 45},
            'PEDRA':        {'P1': 37, 'P2': 40, 'P3': 43, 'P4': 46}
        },
        3: {
            'SALDOS':       {'P1': 49, 'P2': 52, 'P3': 55, 'P4': 58},
            'RECEBIMENTOS': {'P1': 50, 'P2': 53, 'P3': 56, 'P4': 59},
            'PEDRA':        {'P1': 51, 'P2': 54, 'P3': 57, 'P4': 60}
        },
        4: {
            'SALDOS':       {'P1': 63, 'P2': 66, 'P3': 69, 'P4': 72},
            'RECEBIMENTOS': {'P1': 64, 'P2': 67, 'P3': 70, 'P4': 73},
            'PEDRA':        {'P1': 65, 'P2': 68, 'P3': 71, 'P4': 74}
        },
    }

    RELATORIO_DETALHE    = load_workbook("previsao_trens/src/DICIONARIOS/MODELO_RELATORIO_DETALHE.xlsx")
     
    DADOS_RELATORIO = CARREGAR_RELATORIO_DETALHE()  # ['PRINCIPAL', 'RUMO', 'TOTAIS', 'VLI', 'MRS']
    DADOS_RELATORIO = DADOS_RELATORIO["PRINCIPAL"]  # ['TGG', 'TEG', 'CUTRALE', 'T39', 'TES', 'MOEGA X', 'TGRAO', 'CLI', 'TAC', 'T12A', 'BRACELL', 'SBR', 'ECOPORTO', 'HIDROVIAS', 'TEAG']
    
    PAGINA_DETALHE = RELATORIO_DETALHE['DETALHE']
        
    with open(f"previsao_trens/src/DICIONARIOS/MAPA_TERMINAIS_DETALHE.json") as ARQUIVO:
        map_DETALHE = json.load(ARQUIVO)    
    
    for TERMINAL in DADOS_RELATORIO: 
        
        FERROVIAS = list(DADOS_RELATORIO[TERMINAL].keys())
        FERROVIAS.remove('MARGEM')
        FERROVIAS.remove('TOTAL')   
       

        for FERROVIA in FERROVIAS:
            PRODUTOS = list(DADOS_RELATORIO[TERMINAL][FERROVIA].keys()) 

            for PRODUTO in PRODUTOS:
                for DIA in [0, 1, 2, 3, 4]: 
                    for TITULO_COL in list(DADOS_RELATORIO[TERMINAL][FERROVIA][PRODUTO][DIA]):                       
                        if not TITULO_COL == "TT_OF" and  not TITULO_COL == "TT_PD":                             
                            for PERIODO in DADOS_RELATORIO[TERMINAL][FERROVIA][PRODUTO][DIA][TITULO_COL]:

                                LINHA   =   map_DETALHE[TERMINAL][FERROVIA][PRODUTO]
                                COLUNA  =   COLUNAS[DIA][TITULO_COL][PERIODO]
                                VALOR   =   DADOS_RELATORIO[TERMINAL][FERROVIA][PRODUTO][DIA][TITULO_COL][PERIODO]

                                PAGINA_DETALHE.cell(row = LINHA, column= COLUNA,  value = VALOR )


    caminho_static  = os.path.join(settings.BASE_DIR, 'static', 'downloads')
    caminho_arquivo = os.path.join(caminho_static, "RELATORIO_DETALHE.xlsx")
    RELATORIO_DETALHE.save(caminho_arquivo)

    return {"success": True, "mensagem": "planilha baixada"}